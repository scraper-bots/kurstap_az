import asyncio
import aiohttp
from bs4 import BeautifulSoup
import csv
import json
import re
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

class KurstapAsyncScraper:
    def __init__(self, max_concurrent_requests: int = 20):
        self.base_url = "https://www.kurstap.az"
        self.listings_url = f"{self.base_url}/kateqoriyalar"
        self.courses_data = []
        self.max_concurrent_requests = max_concurrent_requests
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

    async def fetch_page(self, session: aiohttp.ClientSession, url: str, params: Dict = None) -> Optional[str]:
        """Fetch a single page asynchronously"""
        try:
            async with session.get(url, params=params, timeout=aiohttp.ClientTimeout(total=30)) as response:
                response.raise_for_status()
                return await response.text()
        except Exception as e:
            print(f"Error fetching {url}: {e}")
            return None

    async def get_course_links_from_page(self, session: aiohttp.ClientSession, offset: int = 0, max_per_page: int = 8) -> List[str]:
        """Extract all course links from a listings page"""
        params = {
            'c': '',
            'index': 'index',
            'vip': '',
            'city': '',
            'underground': '',
            'search': '',
            'company': '',
            'category': '',
            'subCategory': '',
            'subCatTitle': '',
            'title': '',
            'offset': offset,
            'max': max_per_page
        }

        print(f"Fetching listings page (offset={offset})...")
        html = await self.fetch_page(session, self.listings_url, params)

        if not html:
            return []

        soup = BeautifulSoup(html, 'html.parser')

        # Find all course links
        course_links = []
        for link in soup.select('a[href*="/kurslar/"]'):
            href = link.get('href')
            if href and '/kurslar/' in href:
                full_url = f"{self.base_url}{href}" if href.startswith('/') else href
                if full_url not in course_links:
                    course_links.append(full_url)

        print(f"Found {len(course_links)} course links on page (offset={offset})")
        return course_links

    def extract_phone_numbers(self, phone_string: str) -> List[str]:
        """
        Extract individual phone numbers from a concatenated string.
        Handles both formats:
        - +994 XX XXX XX XX (with spaces)
        - +994XXXXXXXXX (without spaces)
        """
        if not phone_string or phone_string.strip() == '':
            return []

        # Pattern to match Azerbaijani phone numbers
        pattern = r'\+994[\s\d]+'
        matches = re.findall(pattern, phone_string)

        # Clean up each phone number (remove extra spaces, normalize)
        cleaned_numbers = []
        for match in matches:
            # Remove any extra whitespace
            cleaned = ' '.join(match.split())
            if cleaned and cleaned not in cleaned_numbers:
                cleaned_numbers.append(cleaned)

        return cleaned_numbers

    async def extract_course_data(self, session: aiohttp.ClientSession, course_url: str, index: int, total: int) -> Optional[List[Dict]]:
        """
        Extract all relevant data from a course page.
        Returns a list of dictionaries - one for each phone number found.
        If no phone numbers, returns a list with one entry.
        """
        try:
            print(f"[{index}/{total}] Scraping: {course_url}")
            html = await self.fetch_page(session, course_url)

            if not html:
                return None

            soup = BeautifulSoup(html, 'html.parser')

            # Base course data (same for all rows)
            base_data = {
                'url': course_url,
                'course_id': course_url.split('/kurslar/')[-1].split('/')[0] if '/kurslar/' in course_url else '',
            }

            # Extract from the course-top-part section
            course_section = soup.select_one('section.course-top-part')
            if not course_section:
                print(f"Warning: Could not find course-top-part section on {course_url}")
                return None

            # Company/Institution name
            main_name = course_section.select_one('a.main-name span:last-child')
            base_data['institution_name'] = main_name.get_text(strip=True) if main_name else ''

            # Course title
            title_desc = course_section.select_one('.title-desc')
            base_data['course_title'] = title_desc.get_text(strip=True) if title_desc else ''

            # Course duration
            duration_elem = course_section.find('span', string=re.compile('Kurs müddəti'))
            if duration_elem:
                duration_p = duration_elem.find_next('p')
                base_data['duration'] = duration_p.get_text(strip=True) if duration_p else ''
            else:
                base_data['duration'] = ''

            # Course price (Fərdi hazırlıq)
            price_elem = course_section.find('span', string=re.compile('Fərdi hazırlıq'))
            if price_elem:
                price_p = price_elem.find_next('p')
                base_data['price'] = price_p.get_text(strip=True) if price_p else ''
            else:
                base_data['price'] = ''

            # City and District
            city_elem = course_section.find('span', string=re.compile('Şəhər, Rayon'))
            if city_elem:
                city_p = city_elem.find_next('p')
                base_data['location'] = city_p.get_text(strip=True).replace('\n', ', ') if city_p else ''
            else:
                base_data['location'] = ''

            # Contact information (phone numbers and email)
            contact_elem = course_section.find('span', string=re.compile('Əlaqə'))
            phone_numbers_raw = []
            emails = []

            if contact_elem:
                contact_ul = contact_elem.find_next('ul')
                if contact_ul:
                    for li in contact_ul.find_all('li'):
                        text = li.get_text(strip=True)
                        # Check if it's a phone number
                        if '+994' in text or any(char.isdigit() for char in text):
                            # Extract individual phone numbers from potentially concatenated string
                            extracted_phones = self.extract_phone_numbers(text)
                            phone_numbers_raw.extend(extracted_phones)
                        # Check if it's an email
                        elif '@' in text:
                            emails.append(text)

            # Store emails as joined string (same for all rows)
            base_data['emails'] = ' | '.join(emails) if emails else ''

            # Address
            address_elem = course_section.find('span', string=re.compile('Ünvan'))
            if address_elem:
                address_p = address_elem.find_next('p')
                base_data['address'] = address_p.get_text(strip=True) if address_p else ''
            else:
                base_data['address'] = ''

            # Social media / Website
            social_elem = course_section.find('span', string=re.compile('Sosial media'))
            website = ''
            if social_elem:
                social_ul = social_elem.find_next('ul')
                if social_ul:
                    link = social_ul.find('a')
                    if link:
                        website = link.get_text(strip=True)
            base_data['website'] = website

            # Create separate row for each phone number
            results = []
            if phone_numbers_raw:
                for phone in phone_numbers_raw:
                    row = base_data.copy()
                    row['phone_numbers'] = phone
                    results.append(row)
            else:
                # No phone numbers found, still add the course data
                row = base_data.copy()
                row['phone_numbers'] = ''
                results.append(row)

            course_title = base_data.get('course_title', 'Unknown')
            num_phones = len(results)
            print(f"✓ [{index}/{total}] Successfully scraped: {course_title} ({num_phones} phone number(s))")
            return results

        except Exception as e:
            print(f"Error extracting data from {course_url}: {e}")
            return None

    async def collect_all_course_urls(self, session: aiohttp.ClientSession) -> List[str]:
        """Collect all course URLs from all pagination pages"""
        print("Collecting all course URLs from listings...")
        offset = 0
        max_per_page = 8
        all_course_urls = set()

        while True:
            course_links = await self.get_course_links_from_page(session, offset, max_per_page)

            if not course_links:
                print(f"No more courses found at offset {offset}. Stopping pagination.")
                break

            all_course_urls.update(course_links)
            offset += max_per_page

        print(f"\n{'='*60}")
        print(f"Total unique course URLs found: {len(all_course_urls)}")
        print(f"{'='*60}\n")

        return list(all_course_urls)

    async def scrape_all_courses(self):
        """Main method to scrape all courses using async/await"""
        print("Starting async scrape...")
        print(f"Max concurrent requests: {self.max_concurrent_requests}\n")

        connector = aiohttp.TCPConnector(limit=self.max_concurrent_requests)
        async with aiohttp.ClientSession(headers=self.headers, connector=connector) as session:
            # First, collect all course URLs
            all_course_urls = await self.collect_all_course_urls(session)

            if not all_course_urls:
                print("No courses found!")
                return

            # Create semaphore to limit concurrent requests
            semaphore = asyncio.Semaphore(self.max_concurrent_requests)

            async def scrape_with_semaphore(url: str, idx: int) -> Optional[List[Dict]]:
                async with semaphore:
                    return await self.extract_course_data(session, url, idx, len(all_course_urls))

            # Scrape all courses concurrently
            print("Starting to scrape individual course pages...\n")
            tasks = [scrape_with_semaphore(url, idx) for idx, url in enumerate(all_course_urls, 1)]
            results = await asyncio.gather(*tasks)

            # Flatten results (since each result is now a list of dicts) and filter out None
            self.courses_data = []
            for result in results:
                if result is not None:
                    self.courses_data.extend(result)  # extend instead of append

        print(f"\n{'='*60}")
        print(f"Scraping complete! Total courses scraped: {len(self.courses_data)}")
        print(f"{'='*60}")

    def save_to_csv(self, filename: str = 'kurstap_courses.csv'):
        """Save scraped data to CSV file"""
        if not self.courses_data:
            print("No data to save!")
            return

        keys = self.courses_data[0].keys()

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=keys)
            writer.writeheader()
            writer.writerows(self.courses_data)

        print(f"✓ Data saved to {filename}")

    def save_to_json(self, filename: str = 'kurstap_courses.json'):
        """Save scraped data to JSON file"""
        if not self.courses_data:
            print("No data to save!")
            return

        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(self.courses_data, f, ensure_ascii=False, indent=2)

        print(f"✓ Data saved to {filename}")

    def save_to_xlsx(self, filename: str = 'kurstap_courses.xlsx'):
        """Save scraped data to Excel file with formatting"""
        if not self.courses_data:
            print("No data to save!")
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Kurstap Courses"

        # Get headers
        headers = list(self.courses_data[0].keys())

        # Style for headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data
        for row_num, course in enumerate(self.courses_data, 2):
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=course.get(header, ''))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Auto-adjust column widths
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            # Set minimum width
            max_length = len(header)
            for row in range(2, min(100, len(self.courses_data) + 2)):  # Check first 100 rows for width
                cell_value = ws.cell(row=row, column=col_num).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))

            # Set width with some padding, but cap at 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = "A2"

        wb.save(filename)
        print(f"✓ Data saved to {filename}")


async def main():
    # Create scraper with max 20 concurrent requests
    scraper = KurstapAsyncScraper(max_concurrent_requests=20)

    # Scrape all courses
    await scraper.scrape_all_courses()

    # Save to CSV, JSON, and XLSX
    print("\nSaving data to files...")
    scraper.save_to_csv('kurstap_courses.csv')
    scraper.save_to_json('kurstap_courses.json')
    scraper.save_to_xlsx('kurstap_courses.xlsx')

    # Print summary
    if scraper.courses_data:
        print(f"\nSample of first course:")
        print(json.dumps(scraper.courses_data[0], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    asyncio.run(main())
